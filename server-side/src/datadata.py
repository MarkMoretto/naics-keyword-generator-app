# server-side\src\datadata.py

import re
from io import BytesIO
import concurrent.futures as ccf



import requests
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


naics_year: int = 2017

urlmap = dict(
    code_index = f"https://www.census.gov/naics/{naics_year}NAICS/{naics_year}_NAICS_Index_File.xlsx",
    code_description = f"https://www.census.gov/naics/{naics_year}NAICS/{naics_year}_NAICS_Descriptions.xlsx"
)




def get_stopwords():
    _url = "https://gist.github.com/sebleier/554280/raw/7e0e4a1ce04c2bb7bd41089c9821dbcf6d0c786c/NLTK's%2520list%2520of%2520english%2520stopwords"
    with requests.Session() as s:
        resp = s.get(_url)
        stopwords_raw = resp.text
        return re.split(r"\n", stopwords_raw)


# --- Retrieve data from census.gov --- #
def retrieve_data(url: str) -> BytesIO:
    """Return BytesIO object representing MS Excel workbook."""
    _resp = requests.get(url)
    _bio = BytesIO(_resp.content)
    if _bio:
        _bio.seek(0)
        return _bio


def data_to_ws(bytedata: BytesIO) -> Worksheet:
    """Return openpyxl Worksheet object."""
    _wb = load_workbook(bytedata)
    return _wb[_wb.sheetnames[0]]


def data_to_df(bytedata: BytesIO) -> pd.DataFrame:
    """Return pd.DataFrame object."""
    return pd.read_excel(bytedata)

# --- Retrieve both assets simultaneously --- #
with ccf.ThreadPoolExecutor(max_workers = len(urlmap) * 2) as e:
    url_futures = {e.submit(retrieve_data, url):url_name for url_name, url in urlmap.items()}
    for fut in ccf.as_completed(url_futures):
        if url_futures[fut] == "code_index":
            df_code = data_to_df(fut.result())
        else:
            df_desc = data_to_df(fut.result())





# List of rows
# Usd for testing only.
rlist = list(ws.rows)




stopwords = get_stopwords()
colmap = {}
for r in rlist[0]:
    colmap[r.column_letter] = r.value
    # ddict[r.column_letter] = []

# # Populate dictionary with data
# for r in rlist[1:101]:
#     for c in r:
#         ddict[c.column_letter].append(c.value)

no_stopwords = lambda s: [w for w in s if not w in stopwords]


clean_pattern = r"[^\w\s\d]"
p = re.compile(clean_pattern, flags = re.I)

ddict = {}
# for r in rlist[1:10]:
for r_idx, r in enumerate(ws.rows, start = 1):
    if r_idx > 11:
        break
    if r_idx > 1:
        __tokens = (
            re.split(r"\s+", \
                p.sub("", str(ws.cell(r_idx, 2).value).lower())
                )
            )
        ddict[ws.cell(r_idx, 1).value] = no_stopwords(__tokens)

        for c_idx, c in enumerate(ws.columns, start = 1):
            print(r_idx, c_idx, ws.cell(r_idx, c_idx).value)
    
sample = "Oilseed farming (except soybean), field and seed production"
re.sub(r"[^\w\s\d]", "", sample)

    for c in r:
        ddict[c.column_letter].append(c.value)


# --- Neural Net / Word2Vec --- #
# https://nathanrooy.github.io/posts/2018-03-22/word2vec-from-scratch-with-python-and-numpy/

class TooWordToVec:
    def __init__(self):
        pass
    

    @staticmethod
    def softmax(x: np.array):
        numer = np.exp(x - np.max(x))
        return numer / numer.sum(axis = 0)

